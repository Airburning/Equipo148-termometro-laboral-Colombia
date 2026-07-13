"""Extrae la serie nacional de 'Proporción de informalidad' (empleo informal)
desde el anexo GEIH-ANEXOS más reciente dentro del zip. La hoja 'Prop
informalidad' viene en trimestres móviles (ventanas de 3 meses) agrupados
en bloques de 12 columnas por año. Cada trimestre se asigna a un semestre
(I = ene-jun, II = jul-dic) según el mes central de la ventana, y luego se
promedian los trimestres de cada semestre para alinear la granularidad con
el panel semestral de GEIH regional.
"""
import zipfile
import io

import openpyxl
import pandas as pd

from config import RAW_DIR, INTERMEDIATE_DIR

ZIP_PATH = RAW_DIR / "GEIH-ANEXOS.zip"
LATEST_ENTRY = "GEIH-ANEXOS/2026/anex-GEIHEISS-feb-abr2026.xlsx"
OUT = INTERMEDIATE_DIR / "informalidad_laboral_nacional.csv"

# posición (0-indexada) dentro del bloque de 12 columnas -> (offset_anio, periodo)
POS_MAP = [
    (0, "I"), (0, "I"), (0, "I"), (0, "I"), (0, "I"),
    (0, "II"), (0, "II"), (0, "II"), (0, "II"), (0, "II"), (0, "II"),
    (1, "I"),
]


def main():
    z = zipfile.ZipFile(ZIP_PATH)
    data = z.read(LATEST_ENTRY)
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb["Prop informalidad"]

    year_row = list(ws.iter_rows(min_row=11, max_row=11, values_only=True))[0]
    data_row = None
    for r in ws.iter_rows(min_row=13, max_row=20, values_only=True):
        if r[0] and str(r[0]).strip().lower() == "total nacional":
            data_row = r
            break
    if data_row is None:
        raise RuntimeError("No se encontró la fila 'Total nacional'")

    # forward-fill del año por columna, y posición dentro del bloque de 12
    col_year = {}
    current_year = None
    block_start = 1
    for c in range(1, len(year_row)):
        if year_row[c] is not None:
            current_year = int(year_row[c])
            block_start = c
        col_year[c] = (current_year, c - block_start)

    records = []
    for c in range(1, len(data_row)):
        val = data_row[c]
        if val is None or col_year.get(c) is None or col_year[c][0] is None:
            continue
        year, pos = col_year[c]
        if pos >= len(POS_MAP):
            continue
        offset, periodo = POS_MAP[pos]
        records.append((year + offset, periodo, val))

    df = pd.DataFrame(records, columns=["anio", "periodo", "valor"])
    agg = df.groupby(["anio", "periodo"], as_index=False)["valor"].mean()
    agg = agg.rename(columns={"valor": "informalidad_laboral_nacional"})
    agg = agg.sort_values(["anio", "periodo"]).reset_index(drop=True)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    agg.to_csv(OUT, index=False)
    print(f"OK -> {OUT}  filas={len(agg)}")
    print(agg.to_string())


if __name__ == "__main__":
    main()
