"""Extrae un índice nacional anual de informalidad empresarial desde IMIE
(hoja 'C2', por sector). No existe una fila de total nacional con datos en
esa hoja, así que se calcula el promedio simple entre sectores de la
'Incidencia de Informalidad Empresarial Multidimensional' por año, como
proxy del nivel nacional de informalidad empresarial.
"""
import openpyxl
import pandas as pd

from config import RAW_DIR, INTERMEDIATE_DIR

SRC = RAW_DIR / "IMIE2024_Procesado.xlsx"
OUT = INTERMEDIATE_DIR / "informalidad_empresarial_nacional.csv"
INDICADOR = "Incidencia de Informalidad Empresarial Multidimensional"


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb["C2"]
    header = list(ws.iter_rows(min_row=14, max_row=14, values_only=True))[0]
    years = [int(y) for y in header[2:8]]

    rows = []
    for r in ws.iter_rows(min_row=15, values_only=True):
        if r[0] is None or r[1] is None:
            continue
        if not str(r[1]).strip().startswith(INDICADOR):
            continue
        for i, year in enumerate(years):
            val = r[2 + i]
            if val is not None:
                rows.append((year, val))

    df = pd.DataFrame(rows, columns=["anio", "valor"])
    agg = df.groupby("anio", as_index=False)["valor"].mean()
    agg = agg.rename(columns={"valor": "informalidad_empresarial_nacional"})
    agg = agg.sort_values("anio").reset_index(drop=True)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    agg.to_csv(OUT, index=False)
    print(f"OK -> {OUT}")
    print(agg.to_string())


if __name__ == "__main__":
    main()
